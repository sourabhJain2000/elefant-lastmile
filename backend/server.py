from fastapi import FastAPI, APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import re
import asyncio
import logging
import urllib.request
from pathlib import Path
from pydantic import BaseModel
from datetime import datetime, timezone, date, timedelta
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1Q7eGyhFSp-GuhqSMFGMvsE-XOQp8o5Hrm1Cwl4vw1YM/edit?gid=1958626721#gid=1958626721"

DELIVERED_STATUSES = {"DELIVERED"}
RETURN_REQUESTED_STATUS = "RETURN_REQUESTED"


# ----------------------------- helpers -----------------------------

def extract_sheet_id(url: str) -> str:
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", url or "")
    if not m:
        raise HTTPException(status_code=400, detail="Invalid Google Sheet URL")
    return m.group(1)


def fetch_workbook(sheet_id: str):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        data = urllib.request.urlopen(req, timeout=120).read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not fetch the Google Sheet. Make sure it is shared as 'Anyone with the link'. ({e})")
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def find_sheet(wb, *keywords):
    for ws in wb.worksheets:
        title = ws.title.lower()
        if all(k.lower() in title for k in keywords):
            return ws
    return None


def sheet_records(ws):
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    header = [str(h).strip() if h is not None else "" for h in header]
    out = []
    for r in rows:
        if r is None:
            continue
        rec = {}
        empty = True
        for h, v in zip(header, r):
            if not h or h in rec:
                continue
            if v is not None and str(v).strip() != "":
                empty = False
            rec[h] = v
        if not empty:
            out.append(rec)
    return out


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("Z", "")
    try:
        return datetime.fromisoformat(s.split(".")[0]).date().isoformat()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).date().isoformat()
        except Exception:
            continue
    return None


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ----------------------------- models -----------------------------

class SyncRequest(BaseModel):
    sheet_url: str | None = None


# ----------------------------- sync -----------------------------

def _parse_workbook(sheet_url: str):
    """Blocking: download + parse the sheet into document lists."""
    sheet_id = extract_sheet_id(sheet_url)
    wb = fetch_workbook(sheet_id)

    orders_ws = find_sheet(wb, "order") or wb.worksheets[0]
    returns_ws = find_sheet(wb, "return")
    serviceable_ws = find_sheet(wb, "serviceable") or find_sheet(wb, "servic")
    library_ws = find_sheet(wb, "librar")

    order_docs = []
    for r in sheet_records(orders_ws):
        status = clean(r.get("Order Status")).upper()
        if status in DELIVERED_STATUSES:
            continue
        order_docs.append({
            "order_id": clean(r.get("Order Id")),
            "product_id": clean(r.get("Product Id")),
            "toy_name": clean(r.get("Toy Name")),
            "library_id": clean(r.get("Library Id")),
            "library_name": clean(r.get("Library Name")) or "Unassigned",
            "user_id": clean(r.get("User Id")),
            "user_name": clean(r.get("User Name")),
            "pincode": clean(r.get("Pincode")),
            "plan_type": clean(r.get("Plan Type")),
            "order_status": status,
            "expected_delivery_date": clean(r.get("Expected Delivery Date")),
            "delivery_date": to_date_str(r.get("Expected Delivery Date")),
            "delivery_type": clean(r.get("Delivery Type")),
            "delivery_partner": clean(r.get("Delivery Partner")),
            "toy_type": clean(r.get("Toy Type")),
            "is_pre_order": clean(r.get("Is Pre ORder")),
            "is_force_order": clean(r.get("Is Force Order")),
            "tags": clean(r.get("Tags")),
        })

    # Returns: store all that are NOT yet RETURN_CONFIRMED (pipeline returns).
    return_docs = []
    for r in sheet_records(returns_ws):
        status = clean(r.get("Return Order Status")).upper()
        if not status or status == "RETURN_CONFIRMED":
            continue
        receiving = clean(r.get("Receiving Library Name"))
        owner = clean(r.get("Owner Library Name"))
        hub = receiving or owner or "Unassigned"
        return_docs.append({
            "return_order": clean(r.get("Return Order")),
            "order_id": clean(r.get("Order Id")),
            "product_id": clean(r.get("Product Id")),
            "product_name": clean(r.get("Product Name")),
            "owner_library": clean(r.get("Owner Library")),
            "owner_library_name": owner,
            "receiving_library": clean(r.get("Receiving Library")),
            "receiving_library_name": receiving,
            "hub_name": hub,
            "user_id": clean(r.get("User Id")),
            "user_name": clean(r.get("User Name")),
            "pincode": clean(r.get("Pincode")),
            "plan_type": clean(r.get("Plan Type")),
            "return_status": status,
            "toy_condition": clean(r.get("Toy Condition")),
            "return_created_at": clean(r.get("Return Created At")),
            "request_date": to_date_str(r.get("Return Created At")),
            "is_pre_order": clean(r.get("Is Pre Order")),
            "is_force_order": clean(r.get("Is Force Order")),
            "tags": clean(r.get("Tags")),
        })

    # Serviceability snapshot (one row per PLACED order).
    serviceable_docs = []
    seen_orders = set()
    for r in sheet_records(serviceable_ws):
        oid = clean(r.get("Order Number")) or clean(r.get("order_id"))
        if not oid or oid in seen_orders:
            continue
        seen_orders.add(oid)
        svc_raw = clean(r.get("Serviceability")) or clean(r.get("Serveable Status"))
        inv = r.get("Available Inventory")
        try:
            inv_num = int(float(inv)) if inv not in (None, "") else 0
        except Exception:
            inv_num = 0
        serviceable_docs.append({
            "order_id": oid,
            "serviceability": svc_raw,
            "order_status": clean(r.get("Order Status")).upper(),
            "order_type": clean(r.get("Order Type")),
            "priority": clean(r.get("Priority Queue")),
            "product_id": clean(r.get("Product Number")),
            "product_name": clean(r.get("Product Name")),
            "library_id": clean(r.get("Library Number")),
            "library_name": clean(r.get("Library Name")) or "Unassigned",
            "city": clean(r.get("City")),
            "user_id": clean(r.get("User Number")),
            "user_name": clean(r.get("User Name")),
            "available_inventory": inv_num,
            "expected_delivery_date": clean(r.get("Expected Delivery Date")),
            "delivery_date": to_date_str(r.get("Expected Delivery Date")),
            "order_placed_on": clean(r.get("Order Placed On")),
        })

    library_docs = []
    for r in sheet_records(library_ws):
        library_docs.append({
            "library_number": clean(r.get("library_number")),
            "name": clean(r.get("name")),
            "lat_long": clean(r.get("lat_long")),
        })

    return order_docs, return_docs, serviceable_docs, library_docs


async def perform_sync(sheet_url: str, trigger: str = "manual"):
    sheet_url = (sheet_url or DEFAULT_SHEET_URL).strip()
    # Run the heavy download + parse off the event loop.
    order_docs, return_docs, serviceable_docs, library_docs = await asyncio.to_thread(_parse_workbook, sheet_url)

    await db.orders.delete_many({})
    await db.returns.delete_many({})
    await db.serviceable.delete_many({})
    await db.libraries.delete_many({})
    if order_docs:
        await db.orders.insert_many(order_docs)
    if return_docs:
        await db.returns.insert_many(return_docs)
    if serviceable_docs:
        await db.serviceable.insert_many(serviceable_docs)
    if library_docs:
        await db.libraries.insert_many(library_docs)

    meta = {
        "_id": "sync",
        "sheet_url": sheet_url,
        "synced_at": datetime.now(timezone.utc).isoformat(),
        "last_trigger": trigger,
        "orders_count": len(order_docs),
        "returns_count": len(return_docs),
        "serviceable_count": len(serviceable_docs),
        "libraries_count": len(library_docs),
    }
    await db.sync_meta.replace_one({"_id": "sync"}, meta, upsert=True)
    meta.pop("_id", None)
    return meta


@api_router.post("/sync")
async def sync_sheet(body: SyncRequest):
    return await perform_sync(body.sheet_url or DEFAULT_SHEET_URL, trigger="manual")


@api_router.get("/sync/status")
async def sync_status():
    meta = await db.sync_meta.find_one({"_id": "sync"}, {"_id": 0})
    if not meta:
        return {"synced": False, "sheet_url": DEFAULT_SHEET_URL}
    meta["synced"] = True
    return meta


# ----------------------------- planning -----------------------------

async def serviceable_map():
    out = {}
    async for d in db.serviceable.find({}, {"_id": 0, "order_id": 1, "serviceability": 1}):
        out[d["order_id"]] = d.get("serviceability", "")
    return out


def parse_plan_date(value):
    if not value:
        return datetime.now(timezone.utc).date()
    try:
        return date.fromisoformat(value)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format, expected YYYY-MM-DD")


async def build_plan(plan_date: date, full: bool = False):
    target_delivery = plan_date + timedelta(days=2)
    return_request_date = plan_date - timedelta(days=2)
    plan_iso = plan_date.isoformat()
    target_iso = target_delivery.isoformat()
    return_iso = return_request_date.isoformat()

    svc = await serviceable_map()

    order_proj = None if full else {
        "_id": 0, "order_id": 1, "toy_name": 1, "toy_type": 1, "user_name": 1,
        "pincode": 1, "order_status": 1, "expected_delivery_date": 1, "tags": 1,
        "delivery_date": 1, "library_name": 1, "library_id": 1,
    }
    return_proj = None if full else {
        "_id": 0, "return_order": 1, "order_id": 1, "product_name": 1, "user_name": 1,
        "pincode": 1, "owner_library_name": 1, "receiving_library_name": 1,
        "return_created_at": 1, "toy_condition": 1, "request_date": 1,
        "hub_name": 1, "receiving_library": 1,
    }

    # Orders to deliver = PLACED/CONFIRMED orders due within the planning horizon
    # (delivery date on or before plan_date + 2 days, which also captures all
    # overdue orders whose delivery date has already passed). De-duplicated to
    # one row per order (an order can have multiple toy/item rows in the sheet).
    order_query = {
        "delivery_date": {"$lte": target_iso, "$ne": None},
        "order_status": {"$in": ["PLACED", "CONFIRMED"]},
    }
    by_order = {}
    async for o in db.orders.find(order_query, order_proj if order_proj else {"_id": 0}):
        oid = o.get("order_id")
        existing = by_order.get(oid)
        if existing is None:
            o["serveable_status"] = svc.get(oid, "")
            o["is_overdue"] = bool(o.get("delivery_date") and o["delivery_date"] < plan_iso)
            o["plan_status"] = "Overdue" if o["is_overdue"] else "Scheduled"
            o["_toys"] = [o.get("toy_name")] if o.get("toy_name") else []
            o["item_count"] = 1
            by_order[oid] = o
        else:
            existing["item_count"] += 1
            if o.get("toy_name"):
                existing["_toys"].append(o.get("toy_name"))
    order_list = list(by_order.values())
    for o in order_list:
        toys = list(dict.fromkeys(o.pop("_toys", [])))
        if toys:
            o["toy_name"] = ", ".join(toys)
    order_list.sort(key=lambda x: (not x["is_overdue"], x.get("delivery_date") or "9999"))

    # Returns (RETURN_REQUESTED): scheduled (requested exactly 2 days ago) +
    # overdue (requested earlier, i.e. request date + 2 days already passed).
    return_list = []
    async for r in db.returns.find({"return_status": "RETURN_REQUESTED", "request_date": {"$lte": return_iso, "$ne": None}}, return_proj if return_proj else {"_id": 0}):
        r["is_overdue"] = bool(r.get("request_date") and r["request_date"] < return_iso)
        r["plan_status"] = "Overdue" if r["is_overdue"] else "Scheduled"
        return_list.append(r)
    return_list.sort(key=lambda x: (not x["is_overdue"], x.get("request_date") or "9999"))

    hubs = {}

    def ensure(name):
        if name not in hubs:
            hubs[name] = {"hub_name": name, "hub_code": "", "orders": [], "returns": []}
        return hubs[name]

    for o in order_list:
        h = ensure(o["library_name"] or "Unassigned")
        if not h["hub_code"] and o.get("library_id"):
            h["hub_code"] = o["library_id"]
        h["orders"].append(o)

    for r in return_list:
        h = ensure(r["hub_name"] or "Unassigned")
        if not h["hub_code"] and r.get("receiving_library"):
            h["hub_code"] = r.get("receiving_library")
        h["returns"].append(r)

    hub_list = []
    for h in hubs.values():
        h["order_count"] = len(h["orders"])
        h["return_count"] = len(h["returns"])
        h["order_overdue_count"] = sum(1 for o in h["orders"] if o.get("is_overdue"))
        h["return_overdue_count"] = sum(1 for r in h["returns"] if r.get("is_overdue"))
        hub_list.append(h)
    hub_list.sort(key=lambda x: (-(x["order_count"] + x["return_count"]), x["hub_name"]))

    return {
        "plan_date": plan_date.isoformat(),
        "target_delivery_date": target_delivery.isoformat(),
        "return_request_date": return_request_date.isoformat(),
        "totals": {
            "hubs": len(hub_list),
            "orders": len(order_list),
            "returns": len(return_list),
            "orders_overdue": sum(1 for o in order_list if o.get("is_overdue")),
            "returns_overdue": sum(1 for r in return_list if r.get("is_overdue")),
        },
        "hubs": hub_list,
    }


@api_router.get("/plan")
async def get_plan(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date)
    return await build_plan(plan_date)


# ----------------------------- order confirmation -----------------------------

# Returns considered "confirmable" — the toy has been physically collected and is
# en route / back at the warehouse, so the return can realistically be confirmed.
CONFIRMABLE_RETURN_STATUSES = {"PICKED_UP", "RETURNED", "ARRIVED"}


async def build_confirmation(plan_date: date | None = None):
    """
    Ready to Confirm — Fully Serviceable PLACED orders (stock available, move
    PLACED -> CONFIRMED and send to WH), grouped by warehouse (serving library).
    When plan_date is given, only orders that need immediate review are kept:
    those whose expected delivery date is on or before plan_date + 2 days.
    """
    target_iso = (plan_date + timedelta(days=2)).isoformat() if plan_date else None

    hubs = {}

    def ensure(name, code):
        if name not in hubs:
            hubs[name] = {"hub_name": name, "hub_code": code or "", "ready_to_confirm": []}
        elif code and not hubs[name]["hub_code"]:
            hubs[name]["hub_code"] = code
        return hubs[name]

    ready_total = 0
    async for s in db.serviceable.find({}, {"_id": 0}):
        if "fully" not in (s.get("serviceability") or "").lower():
            continue
        if target_iso is not None:
            dd = s.get("delivery_date")
            if not dd or dd > target_iso:
                continue
        lib_name = s.get("library_name") or "Unassigned"
        ensure(lib_name, s.get("library_id") or "")["ready_to_confirm"].append({
            "order_id": s.get("order_id"),
            "product_id": s.get("product_id"),
            "product_name": s.get("product_name"),
            "user_name": s.get("user_name"),
            "order_type": s.get("order_type"),
            "available_inventory": s.get("available_inventory", 0),
            "expected_delivery_date": s.get("expected_delivery_date"),
            "order_status": s.get("order_status"),
        })
        ready_total += 1

    hub_list = []
    for h in hubs.values():
        h["ready_count"] = len(h["ready_to_confirm"])
        hub_list.append(h)
    hub_list.sort(key=lambda x: (-x["ready_count"], x["hub_name"]))

    return {
        "totals": {"hubs": len(hub_list), "ready_to_confirm": ready_total},
        "plan_date": plan_date.isoformat() if plan_date else None,
        "target_delivery_date": target_iso,
        "hubs": hub_list,
    }


async def build_return_confirmation():
    """
    Flat list of Not-Serviceable PLACED orders that can be fulfilled by an
    incoming return (PICKED_UP / RETURNED / ARRIVED) of the same product at the
    same warehouse. Return quantity is LIMITED, so it is allocated 1:1 to orders
    prioritised by scheduled delivery date (earliest first), then order number.
    Orders beyond the available return quantity are not listed.
    """
    # Available returns per (product, warehouse). A return restocks its home
    # (owner) library; fall back to the receiving library when owner is blank.
    returns_by_key = {}
    async for r in db.returns.find(
        {"return_status": {"$in": list(CONFIRMABLE_RETURN_STATUSES)}},
        {"_id": 0, "product_id": 1, "owner_library": 1, "receiving_library": 1,
         "return_order": 1, "return_status": 1},
    ):
        pid = r.get("product_id")
        lib = r.get("owner_library") or r.get("receiving_library")
        if not pid or not lib:
            continue
        returns_by_key.setdefault((pid, lib), []).append(
            {"return_order": r.get("return_order"), "return_status": r.get("return_status")}
        )

    # Candidate not-serviceable PLACED orders grouped per (product, warehouse).
    orders_by_key = {}
    async for s in db.serviceable.find({}, {"_id": 0}):
        svc = (s.get("serviceability") or "").lower()
        if "not" not in svc and "partial" not in svc:
            continue
        key = (s.get("product_id"), s.get("library_id"))
        if key not in returns_by_key:
            continue
        orders_by_key.setdefault(key, []).append(s)

    allocated = []
    for key, cand in orders_by_key.items():
        avail = returns_by_key.get(key, [])
        # Prioritise by scheduled delivery date (earliest first), then order no.
        cand.sort(key=lambda s: (s.get("delivery_date") or "9999-12-31", s.get("order_id") or ""))
        for i, s in enumerate(cand):
            if i >= len(avail):
                break  # return quantity exhausted for this product at this WH
            allocated.append({
                "order_id": s.get("order_id"),
                "product_id": s.get("product_id"),
                "product_name": s.get("product_name"),
                "user_name": s.get("user_name"),
                "order_type": s.get("order_type"),
                "library_name": s.get("library_name") or "Unassigned",
                "library_id": s.get("library_id") or "",
                "expected_delivery_date": s.get("expected_delivery_date"),
                "delivery_date": s.get("delivery_date"),
                "matching_returns": [avail[i]],
            })

    allocated.sort(key=lambda x: (x["library_name"], x.get("delivery_date") or "9999-12-31", x["order_id"]))
    return {"totals": {"orders": len(allocated)}, "orders": allocated}


@api_router.get("/confirmation")
async def get_confirmation(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date) if date else None
    return await build_confirmation(plan_date)


@api_router.get("/return-confirmation")
async def get_return_confirmation():
    return await build_return_confirmation()


# Returns that count as potential supply for an order (anything not yet
# confirmed back into inventory and not a failed pickup).
COVERABLE_RETURN_STATUSES = {"RETURN_REQUESTED", "READY_TO_PICKUP", "PICKED_UP", "RETURNED", "ARRIVED"}
# Returns still needing a pickup action.
PICKUP_RETURN_STATUSES = {"RETURN_REQUESTED", "READY_TO_PICKUP"}


async def _not_serviceable_groups():
    """Not-Serviceable PLACED orders grouped by (product_id, library_id)."""
    groups = {}
    async for s in db.serviceable.find({}, {"_id": 0}):
        svc = (s.get("serviceability") or "").lower()
        if "not" not in svc and "partial" not in svc:
            continue
        groups.setdefault((s.get("product_id"), s.get("library_id")), []).append(s)
    for cand in groups.values():
        cand.sort(key=lambda s: (s.get("delivery_date") or "9999-12-31", s.get("order_id") or ""))
    return groups


async def build_unallocatable(plan_date: date | None = None):
    """
    Orders that cannot be allotted any inventory: Not-Serviceable PLACED orders
    that are NOT covered by any usable incoming return (no current stock and no
    return supply). Optional expected-delivery-date filter (can be future) to
    plan ahead. Returns supply is allocated by earliest delivery date first, and
    everything beyond the available supply is unallocatable.
    """
    target_iso = (plan_date + timedelta(days=2)).isoformat() if plan_date else None

    supply = {}
    async for r in db.returns.find(
        {"return_status": {"$in": list(COVERABLE_RETURN_STATUSES)}},
        {"_id": 0, "product_id": 1, "owner_library": 1, "receiving_library": 1},
    ):
        lib = r.get("owner_library") or r.get("receiving_library")
        pid = r.get("product_id")
        if pid and lib:
            supply[(pid, lib)] = supply.get((pid, lib), 0) + 1

    groups = await _not_serviceable_groups()
    out = []
    for key, cand in groups.items():
        sup = supply.get(key, 0)
        for i, s in enumerate(cand):
            if i < sup:
                continue  # covered by available return supply
            dd = s.get("delivery_date")
            if target_iso is not None and (not dd or dd > target_iso):
                continue
            out.append({
                "order_id": s.get("order_id"),
                "product_id": s.get("product_id"),
                "product_name": s.get("product_name"),
                "user_name": s.get("user_name"),
                "order_type": s.get("order_type"),
                "library_name": s.get("library_name") or "Unassigned",
                "library_id": s.get("library_id") or "",
                "available_inventory": s.get("available_inventory", 0),
                "expected_delivery_date": s.get("expected_delivery_date"),
                "delivery_date": dd,
            })
    out.sort(key=lambda x: (x["library_name"], x.get("delivery_date") or "9999-12-31", x["order_id"]))
    return {
        "totals": {"orders": len(out)},
        "plan_date": plan_date.isoformat() if plan_date else None,
        "target_delivery_date": target_iso,
        "orders": out,
    }


async def build_pickup_plan(plan_date: date, window_days: int = 5):
    """
    Returns still needing pickup (RETURN_REQUESTED / READY_TO_PICKUP) that are
    required to fulfil Not-Serviceable orders due within the next `window_days`
    days. Each pending return is allocated to the most urgent order (earliest
    delivery date, then order number) so the team picks up what's needed first.
    """
    cutoff_iso = (plan_date + timedelta(days=window_days)).isoformat()

    pickups = {}
    async for r in db.returns.find(
        {"return_status": {"$in": list(PICKUP_RETURN_STATUSES)}},
        {"_id": 0, "product_id": 1, "owner_library": 1, "receiving_library": 1,
         "return_order": 1, "return_status": 1, "receiving_library_name": 1, "owner_library_name": 1},
    ):
        lib = r.get("owner_library") or r.get("receiving_library")
        pid = r.get("product_id")
        if pid and lib:
            pickups.setdefault((pid, lib), []).append(r)

    groups = await _not_serviceable_groups()
    rows = []
    for key, cand in groups.items():
        avail = pickups.get(key, [])
        if not avail:
            continue
        near = [s for s in cand if s.get("delivery_date") and s["delivery_date"] <= cutoff_iso]
        for i, s in enumerate(near):
            if i >= len(avail):
                break
            r = avail[i]
            rows.append({
                "return_order": r.get("return_order"),
                "return_status": r.get("return_status"),
                "product_name": s.get("product_name"),
                "library_name": s.get("library_name") or "Unassigned",
                "library_id": s.get("library_id") or "",
                "for_order": s.get("order_id"),
                "user_name": s.get("user_name"),
                "order_type": s.get("order_type"),
                "order_delivery_date": s.get("delivery_date"),
                "expected_delivery_date": s.get("expected_delivery_date"),
            })
    rows.sort(key=lambda x: (x.get("order_delivery_date") or "9999-12-31", x["for_order"]))
    return {
        "totals": {"pickups": len(rows)},
        "plan_date": plan_date.isoformat(),
        "window_days": window_days,
        "cutoff_date": cutoff_iso,
        "pickups": rows,
    }


@api_router.get("/unallocatable")
async def get_unallocatable(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date) if date else None
    return await build_unallocatable(plan_date)


@api_router.get("/pickup-plan")
async def get_pickup_plan(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date)
    return await build_pickup_plan(plan_date)



# ----------------------------- excel export -----------------------------

ORDER_COLUMNS = [
    ("order_id", "Order Id"),
    ("plan_status", "Plan Status"),
    ("toy_name", "Toy Name"),
    ("item_count", "Items"),
    ("toy_type", "Toy Type"),
    ("user_name", "User Name"),
    ("pincode", "Pincode"),
    ("order_status", "Order Status"),
    ("expected_delivery_date", "Expected Delivery Date"),
    ("serveable_status", "Serviceable Status"),
    ("is_pre_order", "Is Pre Order"),
    ("is_force_order", "Is Force Order"),
    ("delivery_partner", "Delivery Partner"),
    ("tags", "Tags"),
]

RETURN_COLUMNS = [
    ("return_order", "Return Order"),
    ("plan_status", "Plan Status"),
    ("order_id", "Order Id"),
    ("product_name", "Product Name"),
    ("user_name", "User Name"),
    ("pincode", "Pincode"),
    ("owner_library_name", "Owner Hub"),
    ("receiving_library_name", "Receiving Hub"),
    ("return_created_at", "Return Requested At"),
    ("toy_condition", "Toy Condition"),
    ("tags", "Tags"),
]

HEADER_FILL = openpyxl.styles.PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = openpyxl.styles.Font(color="FFFFFF", bold=True)
SECTION_FONT = openpyxl.styles.Font(bold=True, size=12)


def _safe_sheet_name(name, used):
    invalid = set('[]:*?/\\')
    cleaned = "".join(c for c in name if c not in invalid).strip() or "Hub"
    cleaned = cleaned[:28]
    base = cleaned
    i = 1
    while cleaned.lower() in used:
        cleaned = f"{base[:25]}_{i}"
        i += 1
    used.add(cleaned.lower())
    return cleaned


def _autosize(ws):
    for col in ws.columns:
        length = 0
        letter = None
        for cell in col:
            if letter is None:
                letter = cell.column_letter
            try:
                length = max(length, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = min(max(length + 2, 10), 45)


def _write_hub_sheet(ws, hub, plan):
    row = 1
    ws.cell(row=row, column=1, value=f"{hub['hub_name']}  ({hub.get('hub_code','')})").font = openpyxl.styles.Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Plan Date: {plan['plan_date']}   |   Delivery Date: {plan['target_delivery_date']}   |   Return Requested On: {plan['return_request_date']}")
    row += 2

    ws.cell(row=row, column=1, value=f"ORDERS TO DELIVER ({hub['order_count']})").font = SECTION_FONT
    row += 1
    for ci, (_, label) in enumerate(ORDER_COLUMNS, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row += 1
    for o in hub["orders"]:
        for ci, (key, _) in enumerate(ORDER_COLUMNS, start=1):
            ws.cell(row=row, column=ci, value=o.get(key, ""))
        row += 1
    row += 2

    ws.cell(row=row, column=1, value=f"RETURNS TO PICK UP ({hub['return_count']})").font = SECTION_FONT
    row += 1
    for ci, (_, label) in enumerate(RETURN_COLUMNS, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row += 1
    for r in hub["returns"]:
        for ci, (key, _) in enumerate(RETURN_COLUMNS, start=1):
            ws.cell(row=row, column=ci, value=r.get(key, ""))
        row += 1
    _autosize(ws)


def build_workbook(plan, single_hub=None):
    wb = openpyxl.Workbook()
    used = set()

    hubs = plan["hubs"]
    if single_hub is not None:
        hubs = [h for h in plan["hubs"] if h["hub_name"] == single_hub]
        if not hubs:
            raise HTTPException(status_code=404, detail="Hub not found in plan")

    if single_hub is None:
        ws = wb.active
        ws.title = _safe_sheet_name("Summary", used)
        ws.cell(row=1, column=1, value="LAST MILE DAILY PLAN").font = openpyxl.styles.Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"Plan Date: {plan['plan_date']}")
        ws.cell(row=3, column=1, value=f"Delivery Date (+2 days): {plan['target_delivery_date']}")
        ws.cell(row=4, column=1, value=f"Returns Requested On (-2 days): {plan['return_request_date']}")
        ws.cell(row=6, column=1, value=f"Total Hubs: {plan['totals']['hubs']}   Total Orders: {plan['totals']['orders']}   Total Returns: {plan['totals']['returns']}")
        head_row = 8
        for ci, label in enumerate(["Hub", "Hub Code", "Orders", "Returns", "Total"], start=1):
            c = ws.cell(row=head_row, column=ci, value=label)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        rr = head_row + 1
        for h in plan["hubs"]:
            ws.cell(row=rr, column=1, value=h["hub_name"])
            ws.cell(row=rr, column=2, value=h.get("hub_code", ""))
            ws.cell(row=rr, column=3, value=h["order_count"])
            ws.cell(row=rr, column=4, value=h["return_count"])
            ws.cell(row=rr, column=5, value=h["order_count"] + h["return_count"])
            rr += 1
        _autosize(ws)
        for h in hubs:
            sheet = wb.create_sheet(title=_safe_sheet_name(h["hub_name"], used))
            _write_hub_sheet(sheet, h, plan)
    else:
        h = hubs[0]
        ws = wb.active
        ws.title = _safe_sheet_name(h["hub_name"], used)
        _write_hub_sheet(ws, h, plan)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/plan/export")
async def export_plan(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date)
    plan = await build_plan(plan_date, full=True)
    buf = build_workbook(plan)
    fname = f"last_mile_plan_{plan['plan_date']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/plan/export/hub")
async def export_hub(hub: str, date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date)
    plan = await build_plan(plan_date, full=True)
    buf = build_workbook(plan, single_hub=hub)
    safe = re.sub(r"[^A-Za-z0-9]+", "_", hub).strip("_")
    fname = f"plan_{safe}_{plan['plan_date']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


CONFIRM_READY_COLUMNS = [
    ("order_id", "Order Id"),
    ("product_name", "Product Name"),
    ("product_id", "Product Id"),
    ("user_name", "User Name"),
    ("order_type", "Order Type"),
    ("available_inventory", "Available Inventory"),
    ("expected_delivery_date", "Expected Delivery Date"),
    ("order_status", "Current Status"),
]

CONFIRM_AWAIT_COLUMNS = [
    ("order_id", "Order Id"),
    ("product_name", "Product Name"),
    ("product_id", "Product Id"),
    ("user_name", "User Name"),
    ("order_type", "Order Type"),
    ("library_name", "Warehouse"),
    ("returns_summary", "Pending Returns To Confirm"),
]


def _write_confirmation_sheet(ws, hub):
    row = 1
    ws.cell(row=row, column=1, value=f"{hub['hub_name']}  ({hub.get('hub_code','')})").font = openpyxl.styles.Font(bold=True, size=14)
    row += 2

    ws.cell(row=row, column=1, value=f"READY TO CONFIRM — SEND TO WH ({hub['ready_count']})").font = SECTION_FONT
    row += 1
    for ci, (_, label) in enumerate(CONFIRM_READY_COLUMNS, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row += 1
    for o in hub["ready_to_confirm"]:
        for ci, (key, _) in enumerate(CONFIRM_READY_COLUMNS, start=1):
            ws.cell(row=row, column=ci, value=o.get(key, ""))
        row += 1
    _autosize(ws)


def build_confirmation_workbook(conf, single_hub=None):
    wb = openpyxl.Workbook()
    used = set()
    hubs = conf["hubs"]
    if single_hub is not None:
        hubs = [h for h in conf["hubs"] if h["hub_name"] == single_hub]
        if not hubs:
            raise HTTPException(status_code=404, detail="Hub not found")

    if single_hub is None:
        ws = wb.active
        ws.title = _safe_sheet_name("Summary", used)
        ws.cell(row=1, column=1, value="ORDER CONFIRMATION PLAN").font = openpyxl.styles.Font(bold=True, size=16)
        ws.cell(row=3, column=1, value=f"Total Hubs: {conf['totals']['hubs']}   Ready to Confirm: {conf['totals']['ready_to_confirm']}")
        head_row = 5
        for ci, label in enumerate(["Hub", "Hub Code", "Ready to Confirm"], start=1):
            c = ws.cell(row=head_row, column=ci, value=label)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        rr = head_row + 1
        for h in conf["hubs"]:
            ws.cell(row=rr, column=1, value=h["hub_name"])
            ws.cell(row=rr, column=2, value=h.get("hub_code", ""))
            ws.cell(row=rr, column=3, value=h["ready_count"])
            rr += 1
        _autosize(ws)
        for h in hubs:
            _write_confirmation_sheet(wb.create_sheet(title=_safe_sheet_name(h["hub_name"], used)), h)
    else:
        _write_confirmation_sheet(wb.active, hubs[0])
        wb.active.title = _safe_sheet_name(hubs[0]["hub_name"], used)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_return_confirmation_workbook(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confirm via Returns"
    ws.cell(row=1, column=1, value="ORDERS CONFIRMABLE VIA RETURNS").font = openpyxl.styles.Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=f"Total Orders: {data['totals']['orders']}  (returns in PICKED_UP / RETURNED / ARRIVED status)")
    head_row = 4
    for ci, (_, label) in enumerate(CONFIRM_AWAIT_COLUMNS, start=1):
        c = ws.cell(row=head_row, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    rr = head_row + 1
    for o in data["orders"]:
        summary = "; ".join(
            f"{m.get('return_order')} ({m.get('return_status')})" for m in o.get("matching_returns", [])
        )
        vals = dict(o)
        vals["returns_summary"] = summary
        for ci, (key, _) in enumerate(CONFIRM_AWAIT_COLUMNS, start=1):
            ws.cell(row=rr, column=ci, value=vals.get(key, ""))
        rr += 1
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/confirmation/export")
async def export_confirmation(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date) if date else None
    conf = await build_confirmation(plan_date)
    buf = build_confirmation_workbook(conf)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="order_confirmation_plan.xlsx"'},
    )


@api_router.get("/confirmation/export/hub")
async def export_confirmation_hub(hub: str, date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date) if date else None
    conf = await build_confirmation(plan_date)
    buf = build_confirmation_workbook(conf, single_hub=hub)
    safe = re.sub(r"[^A-Za-z0-9]+", "_", hub).strip("_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="confirmation_{safe}.xlsx"'},
    )


@api_router.get("/return-confirmation/export")
async def export_return_confirmation():
    data = await build_return_confirmation()
    buf = build_return_confirmation_workbook(data)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="confirm_via_returns.xlsx"'},
    )


UNALLOC_COLUMNS = [
    ("order_id", "Order Id"),
    ("product_name", "Product Name"),
    ("product_id", "Product Id"),
    ("user_name", "User Name"),
    ("order_type", "Order Type"),
    ("library_name", "Warehouse"),
    ("available_inventory", "Available Inventory"),
    ("expected_delivery_date", "Expected Delivery Date"),
]

PICKUP_COLUMNS = [
    ("return_order", "Return Order"),
    ("return_status", "Return Status"),
    ("product_name", "Product Name"),
    ("library_name", "Warehouse"),
    ("for_order", "For Order"),
    ("user_name", "Customer"),
    ("expected_delivery_date", "Order Delivery Date"),
]


def _flat_workbook(title, header, columns, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=title).font = openpyxl.styles.Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=header)
    hr = 4
    for ci, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=hr, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    rr = hr + 1
    for o in rows:
        for ci, (key, _) in enumerate(columns, start=1):
            ws.cell(row=rr, column=ci, value=o.get(key, ""))
        rr += 1
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/unallocatable/export")
async def export_unallocatable(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date) if date else None
    data = await build_unallocatable(plan_date)
    hdr = f"Total: {data['totals']['orders']}" + (f"  |  Expected delivery on or before {data['target_delivery_date']}" if data.get('target_delivery_date') else "  |  All dates")
    buf = _flat_workbook("ORDERS WITH NO INVENTORY (UNALLOCATABLE)", hdr, UNALLOC_COLUMNS, data["orders"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="unallocatable_orders.xlsx"'},
    )


@api_router.get("/pickup-plan/export")
async def export_pickup_plan(date: str | None = Query(default=None)):
    plan_date = parse_plan_date(date)
    data = await build_pickup_plan(plan_date)
    hdr = f"Total returns to pick up: {data['totals']['pickups']}  |  For orders delivering on or before {data['cutoff_date']} (next {data['window_days']} days)"
    buf = _flat_workbook("RETURN PICKUP PLAN", hdr, PICKUP_COLUMNS, data["pickups"])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="return_pickup_plan.xlsx"'},
    )




@api_router.get("/")
async def root():
    return {"message": "Last Mile Planner API"}


AUTO_SYNC_INTERVAL_SECONDS = 15 * 60


async def auto_sync_loop():
    """Background task: re-sync the sheet every 15 minutes."""
    while True:
        await asyncio.sleep(AUTO_SYNC_INTERVAL_SECONDS)
        try:
            meta = await db.sync_meta.find_one({"_id": "sync"}, {"_id": 0, "sheet_url": 1})
            sheet_url = (meta or {}).get("sheet_url") or DEFAULT_SHEET_URL
            result = await perform_sync(sheet_url, trigger="auto")
            logger.info("Auto-sync OK: %s orders, %s returns", result["orders_count"], result["returns_count"])
        except asyncio.CancelledError:
            raise
        except Exception as e:
            logger.error("Auto-sync failed: %s", e)


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def start_auto_sync():
    app.state.auto_sync_task = asyncio.create_task(auto_sync_loop())
    logger.info("Auto-sync scheduled every %s minutes", AUTO_SYNC_INTERVAL_SECONDS // 60)


@app.on_event("shutdown")
async def shutdown_db_client():
    task = getattr(app.state, "auto_sync_task", None)
    if task:
        task.cancel()
    client.close()
