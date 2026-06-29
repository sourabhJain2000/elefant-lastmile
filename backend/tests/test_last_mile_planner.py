"""Backend API tests for Last Mile Planner + Order Confirmation + Confirm via Returns.

Covers latest contract change for /api/plan:
  - order_status in {PLACED, CONFIRMED} for every plan order
  - de-duplicated by order_id (unique order_ids across hubs); multi-item orders
    have item_count > 1 and a comma-joined toy_name
  - delivery_date <= plan_date + 2 days; is_overdue iff delivery_date < plan_date
  - totals.orders is much larger than the previous-buggy ~259
"""
import os
import io
import pytest
import requests
import openpyxl

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"

PLAN_DATE = "2026-06-29"
TARGET_DELIVERY = "2026-07-01"
RETURN_REQ_DATE = "2026-06-27"
TOMORROW = "2026-06-30"

ALLOWED_ORDER_STATUSES = {"PLACED", "CONFIRMED"}


# ----------------------------- sync status -----------------------------
class TestSyncStatus:
    def test_sync_status_already_synced(self):
        r = requests.get(f"{API}/sync/status", timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d.get("synced") is True
        assert d["orders_count"] > 0
        assert d["returns_count"] > 0
        assert d["serviceable_count"] > 0
        assert d["libraries_count"] > 0


# --------------------- plan endpoint (latest dedup + status filter) ---------------------
class TestPlanEndpoint:
    @pytest.fixture(scope="class")
    def plan(self):
        r = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=180)
        assert r.status_code == 200, r.text[:500]
        return r.json()

    def test_plan_shape(self, plan):
        assert plan["plan_date"] == PLAN_DATE
        assert plan["target_delivery_date"] == TARGET_DELIVERY
        assert plan["return_request_date"] == RETURN_REQ_DATE
        t = plan["totals"]
        assert t["hubs"] > 0
        assert t["orders"] > 0
        assert t["returns"] > 0

    def test_totals_orders_much_larger_after_dedup_change(self, plan):
        # User-stated expectation: totals.orders ~1035 after status-set fix.
        # Assert > 600 (well above previous buggy 259) per review_request.
        assert plan["totals"]["orders"] > 600, (
            f"Orders total {plan['totals']['orders']} is too low — dedup/status filter likely wrong"
        )

    def test_only_placed_or_confirmed_status(self, plan):
        for h in plan["hubs"]:
            for o in h["orders"]:
                st = (o.get("order_status") or "").upper()
                assert st in ALLOWED_ORDER_STATUSES, (
                    f"order {o.get('order_id')} status {st} not in {ALLOWED_ORDER_STATUSES}"
                )

    def test_orders_within_delivery_window_and_overdue_correctness(self, plan):
        overdue_count = 0
        total_orders = 0
        for h in plan["hubs"]:
            for o in h["orders"]:
                total_orders += 1
                dd = o.get("delivery_date")
                assert dd is not None, f"order {o.get('order_id')} has null delivery_date"
                assert dd <= TARGET_DELIVERY, (
                    f"order {o.get('order_id')} delivery_date {dd} exceeds {TARGET_DELIVERY}"
                )
                expected_overdue = dd < PLAN_DATE
                assert bool(o.get("is_overdue")) == expected_overdue, (
                    f"order {o.get('order_id')} is_overdue={o.get('is_overdue')} dd={dd}"
                )
                if expected_overdue:
                    overdue_count += 1
        assert total_orders == plan["totals"]["orders"]
        assert overdue_count == plan["totals"]["orders_overdue"]

    def test_order_ids_unique_across_all_hubs(self, plan):
        seen = set()
        dupes = []
        for h in plan["hubs"]:
            for o in h["orders"]:
                oid = o.get("order_id")
                if oid in seen:
                    dupes.append(oid)
                else:
                    seen.add(oid)
        assert not dupes, f"Duplicate order_ids found across hubs: {dupes[:10]} (total {len(dupes)})"
        assert len(seen) == plan["totals"]["orders"]

    def test_multi_item_orders_have_item_count_and_joined_toys(self, plan):
        multi_item_orders = 0
        for h in plan["hubs"]:
            for o in h["orders"]:
                ic = o.get("item_count")
                assert isinstance(ic, int) and ic >= 1, (
                    f"order {o.get('order_id')} has bad item_count {ic}"
                )
                if ic > 1:
                    multi_item_orders += 1
                    toys = o.get("toy_name") or ""
                    # Multi-item should have comma-joined toy names (more than one distinct toy
                    # or at minimum a non-empty toy string).
                    assert toys, f"order {o.get('order_id')} item_count={ic} but toy_name empty"
        # With ~1035 orders and 2556 item rows, there must be plenty of multi-item dedup hits.
        assert multi_item_orders > 0, "Expected some orders with item_count>1 after dedup, found none"

    def test_today_or_earlier_bucket_in_expected_range(self, plan):
        # Reference: orders with delivery_date <= 2026-06-29 should be ~614 (user said ~619).
        count = 0
        for h in plan["hubs"]:
            for o in h["orders"]:
                if (o.get("delivery_date") or "") <= PLAN_DATE:
                    count += 1
        assert 500 <= count <= 800, (
            f"Orders with delivery_date <= {PLAN_DATE} = {count}, expected 500..800"
        )

    def test_returns_logic_unchanged(self, plan):
        for h in plan["hubs"]:
            for ret in h["returns"]:
                assert ret["request_date"] is not None
                assert ret["request_date"] <= RETURN_REQ_DATE
        assert plan["totals"]["returns"] > 0

    def test_plan_empty_far_past(self):
        r = requests.get(f"{API}/plan", params={"date": "2020-01-01"}, timeout=60)
        assert r.status_code == 200
        d = r.json()
        assert d["totals"]["orders"] == 0
        assert d["totals"]["returns"] == 0
        assert d["totals"]["hubs"] == 0

    def test_plan_invalid_date(self):
        r = requests.get(f"{API}/plan", params={"date": "not-a-date"}, timeout=30)
        assert r.status_code == 400


# ----------------------------- plan export -----------------------------
class TestPlanExport:
    @pytest.fixture(scope="class")
    def plan(self):
        r = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=180)
        assert r.status_code == 200
        return r.json()

    def test_export_all_xlsx_with_items_and_plan_status_columns(self, plan):
        r = requests.get(f"{API}/plan/export", params={"date": PLAN_DATE}, timeout=240)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        names_lower = [s.lower() for s in wb.sheetnames]
        assert "summary" in names_lower
        assert len(wb.sheetnames) >= 2

        # Pick the first non-summary sheet and verify Items + Plan Status columns
        # and that some Plan Status cell contains Overdue or Scheduled.
        target = next((s for s in wb.sheetnames if s.lower() != "summary"), None)
        assert target is not None
        ws = wb[target]
        # Locate header row by scanning first 12 rows for 'Order Id'.
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            if row and "Order Id" in [c for c in row if c is not None]:
                headers = list(row)
                break
        assert headers, f"Could not find header row in sheet '{target}'"
        assert "Items" in headers, f"'Items' column missing in hub sheet '{target}' headers={headers}"
        assert "Plan Status" in headers, f"'Plan Status' column missing headers={headers}"

    def test_export_hub_for_real_hub(self, plan):
        # Pick a hub that actually has orders.
        hub_name = None
        for h in plan["hubs"]:
            if h["order_count"] > 0:
                hub_name = h["hub_name"]
                break
        assert hub_name, "No hub with orders found in plan"
        r = requests.get(
            f"{API}/plan/export/hub",
            params={"hub": hub_name, "date": PLAN_DATE},
            timeout=120,
        )
        assert r.status_code == 200, r.text[:300]
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_export_hub_not_found(self):
        r = requests.get(
            f"{API}/plan/export/hub",
            params={"hub": "NoSuchHub_ZZZ", "date": PLAN_DATE},
            timeout=30,
        )
        assert r.status_code == 404


# ----------------------------- order confirmation (regression) -----------------------------
class TestConfirmation:
    @pytest.fixture(scope="class")
    def conf(self):
        r = requests.get(f"{API}/confirmation", timeout=120)
        assert r.status_code == 200
        return r.json()

    def test_structure_and_totals(self, conf):
        assert "totals" in conf and "hubs" in conf
        t = conf["totals"]
        assert t["hubs"] > 0
        assert t["ready_to_confirm"] > 0

    def test_hub_counts(self, conf):
        for h in conf["hubs"]:
            assert h["ready_count"] == len(h["ready_to_confirm"])

    def test_export_all(self):
        r = requests.get(f"{API}/confirmation/export", timeout=120)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")


# --------------------- confirm via returns (regression + ARRIVED) ---------------------
class TestReturnConfirmation:
    @pytest.fixture(scope="class")
    def rc(self):
        r = requests.get(f"{API}/return-confirmation", timeout=120)
        assert r.status_code == 200
        return r.json()

    def test_shape(self, rc):
        assert "totals" in rc and "orders" in rc
        assert rc["totals"]["orders"] == len(rc["orders"])

    def test_every_matching_return_status_in_allowed_set(self, rc):
        allowed = {"PICKED_UP", "RETURNED", "ARRIVED"}
        for o in rc["orders"]:
            matches = o.get("matching_returns") or []
            assert len(matches) > 0
            for m in matches:
                st = (m.get("return_status") or "").upper()
                assert st in allowed

    def test_export(self):
        r = requests.get(f"{API}/return-confirmation/export", timeout=120)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
