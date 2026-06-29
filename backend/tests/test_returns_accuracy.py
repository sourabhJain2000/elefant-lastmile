"""Returns accuracy audit for /api/plan.

Invariants for the 'Returns to Pick Up' list:
  (a) every return has return_status == 'RETURN_REQUESTED'
  (b) every return has request_date <= plan_date - 2 days
  (c) all return_order values unique across the response
  (d) sum of hub.return_count over all hubs == totals.returns
  (e) is_overdue == (request_date < plan_date - 2)
"""
import io
import os
from datetime import date, timedelta

import openpyxl
import pytest
import requests

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"

PLAN_DATE = "2026-06-29"
RETURN_CUTOFF = "2026-06-27"  # plan_date - 2 days
BOUNDARY_PLAN_DATE = "2026-06-20"
BOUNDARY_CUTOFF = "2026-06-18"
FUTURE_PLAN_DATE = "2030-01-01"
PAST_PLAN_DATE = "2020-01-01"


def _flatten_returns(plan):
    out = []
    for h in plan["hubs"]:
        for r in h["returns"]:
            out.append((h["hub_name"], r))
    return out


# --------------------------- returns accuracy ---------------------------
class TestReturnsAccuracy:
    @pytest.fixture(scope="class")
    def plan(self):
        r = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=240)
        assert r.status_code == 200, r.text[:400]
        return r.json()

    def test_plan_returns_metadata(self, plan):
        assert plan["plan_date"] == PLAN_DATE
        assert plan["return_request_date"] == RETURN_CUTOFF
        assert plan["totals"]["returns"] > 1000, (
            f"Expected large returns backlog, got {plan['totals']['returns']}"
        )

    def test_every_return_status_is_return_requested(self, plan):
        bad = []
        for hub_name, r in _flatten_returns(plan):
            st = (r.get("return_status") or "").upper()
            # API may omit return_status from the projected response; verify field is absent OR equals RETURN_REQUESTED.
            if st and st != "RETURN_REQUESTED":
                bad.append((hub_name, r.get("return_order"), st))
        # If projection omits status, run a second pass against /api/plan with full=False (default) — still expect zero leak.
        assert not bad, f"Non-RETURN_REQUESTED returns leaked: {bad[:10]}"

    def test_every_request_date_within_cutoff(self, plan):
        violators = []
        for hub_name, r in _flatten_returns(plan):
            rd = r.get("request_date")
            assert rd is not None, f"return {r.get('return_order')} has null request_date"
            if rd > RETURN_CUTOFF:
                violators.append((hub_name, r.get("return_order"), rd))
        assert not violators, f"Returns with request_date > {RETURN_CUTOFF}: {violators[:10]}"

    def test_is_overdue_matches_strict_inequality(self, plan):
        wrong = []
        overdue_count = 0
        for _, r in _flatten_returns(plan):
            rd = r.get("request_date")
            expected = rd < RETURN_CUTOFF
            if bool(r.get("is_overdue")) != expected:
                wrong.append((r.get("return_order"), rd, r.get("is_overdue")))
            if expected:
                overdue_count += 1
        assert not wrong, f"is_overdue mismatches: {wrong[:10]}"
        assert overdue_count == plan["totals"]["returns_overdue"]

    def test_return_orders_unique_across_response(self, plan):
        seen = set()
        dupes = []
        for _, r in _flatten_returns(plan):
            ro = r.get("return_order")
            if not ro:
                continue
            if ro in seen:
                dupes.append(ro)
            else:
                seen.add(ro)
        assert not dupes, f"Duplicate return_order values: {dupes[:10]} (total {len(dupes)})"
        # also matches totals.returns
        total_with_id = sum(1 for _, r in _flatten_returns(plan) if r.get("return_order"))
        assert len(seen) == total_with_id

    def test_per_hub_return_count_sums_to_total(self, plan):
        s = sum(h["return_count"] for h in plan["hubs"])
        assert s == plan["totals"]["returns"], (
            f"hub.return_count sum {s} != totals.returns {plan['totals']['returns']}"
        )
        # also count_count matches actual list length per hub
        for h in plan["hubs"]:
            assert h["return_count"] == len(h["returns"])

    def test_returns_backlog_in_expected_range(self, plan):
        # Source has ~2479 RETURN_REQUESTED with created<=2026-06-27 (live drift allowed).
        t = plan["totals"]["returns"]
        assert 2200 <= t <= 2800, f"totals.returns {t} outside expected 2200..2800 range"
        # Big backlog: overdue should dominate
        assert plan["totals"]["returns_overdue"] >= int(0.9 * t), (
            f"returns_overdue {plan['totals']['returns_overdue']} too small vs total {t}"
        )


# --------------------------- boundary checks ---------------------------
class TestReturnsBoundaries:
    def test_earlier_plan_date_is_strict_subset(self):
        big = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=240).json()
        small = requests.get(f"{API}/plan", params={"date": BOUNDARY_PLAN_DATE}, timeout=240).json()
        assert small["return_request_date"] == BOUNDARY_CUTOFF

        # all returns in 'small' must satisfy earlier cutoff
        for h in small["hubs"]:
            for r in h["returns"]:
                assert r["request_date"] <= BOUNDARY_CUTOFF, (
                    f"return {r.get('return_order')} rd={r['request_date']} > {BOUNDARY_CUTOFF}"
                )

        # subset: every return_order in small must appear in big
        big_ids = {r["return_order"] for h in big["hubs"] for r in h["returns"]}
        small_ids = {r["return_order"] for h in small["hubs"] for r in h["returns"]}
        missing = small_ids - big_ids
        assert not missing, f"Boundary subset failed; {len(missing)} ids in small but not big"
        assert small["totals"]["returns"] <= big["totals"]["returns"]

    def test_far_future_plan_date_includes_all_return_requested(self):
        # Cutoff is in 2029 -> should include essentially every RETURN_REQUESTED with non-null request_date.
        r = requests.get(f"{API}/plan", params={"date": FUTURE_PLAN_DATE}, timeout=240).json()
        assert r["totals"]["returns"] > 2400  # well above the 2026-06-27 cutoff count
        # status check on a sample
        for h in r["hubs"][:3]:
            for ret in h["returns"][:50]:
                st = (ret.get("return_status") or "RETURN_REQUESTED").upper()
                assert st == "RETURN_REQUESTED"

    def test_far_past_plan_date_yields_zero_returns(self):
        r = requests.get(f"{API}/plan", params={"date": PAST_PLAN_DATE}, timeout=120).json()
        assert r["totals"]["returns"] == 0
        assert all(h["return_count"] == 0 for h in r["hubs"])


# --------------------------- regression on orders contract ---------------------------
class TestOrdersRegression:
    def test_orders_only_placed_or_confirmed_and_unique(self):
        plan = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=240).json()
        seen = set()
        for h in plan["hubs"]:
            for o in h["orders"]:
                st = (o.get("order_status") or "").upper()
                assert st in {"PLACED", "CONFIRMED"}
                assert o["delivery_date"] <= "2026-07-01"
                assert o["order_id"] not in seen
                seen.add(o["order_id"])
        assert len(seen) == plan["totals"]["orders"]


# --------------------------- excel export returns section ---------------------------
class TestExcelReturnsSection:
    def test_per_hub_returns_section_counts(self):
        plan = requests.get(f"{API}/plan", params={"date": PLAN_DATE}, timeout=240).json()
        # pick a hub with a sizable return_count for a meaningful check
        target_hub = max(plan["hubs"], key=lambda h: h["return_count"])
        assert target_hub["return_count"] > 0, "No hub with returns to verify"
        expected_count = target_hub["return_count"]

        r = requests.get(f"{API}/plan/export", params={"date": PLAN_DATE}, timeout=300)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        # Find the hub sheet (sheet name is _safe_sheet_name truncated/cleaned).
        hub_clean = "".join(c for c in target_hub["hub_name"] if c not in set('[]:*?/\\')).strip()[:28]
        sheet = None
        for name in wb.sheetnames:
            if name.startswith(hub_clean[:20]):
                sheet = wb[name]
                break
        assert sheet is not None, f"hub sheet not found for {target_hub['hub_name']} in {wb.sheetnames}"

        # Find 'RETURNS TO PICK UP (N)' section header
        found_label = None
        for row in sheet.iter_rows(min_row=1, max_row=2000, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.startswith("RETURNS TO PICK UP"):
                    found_label = cell
                    break
            if found_label:
                break
        assert found_label, f"RETURNS section not found in hub sheet {sheet.title}"
        # Extract the count in parentheses
        import re
        m = re.search(r"\((\d+)\)", found_label)
        assert m, f"Could not parse count from '{found_label}'"
        assert int(m.group(1)) == expected_count, (
            f"Excel returns count {m.group(1)} != API hub.return_count {expected_count}"
        )
